"""
Validation Sampling Module

Creates stratified samples for manual validation and generates
Excel worksheets for human reviewers.
"""

import random
from pathlib import Path
from typing import Optional
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_validation_sample(results: list[dict],
                              sample_size: int = 50,
                              stratify_by: str = 'document_type',
                              include_errors: bool = True,
                              seed: Optional[int] = None) -> list[dict]:
    """
    Create a stratified sample for manual validation.

    Args:
        results: List of extraction result dictionaries
        sample_size: Target number of documents to sample
        stratify_by: Field to stratify by ('document_type' or 'extraction_status')
        include_errors: Whether to oversample error cases
        seed: Random seed for reproducibility

    Returns:
        List of sampled extraction results with sampling metadata
    """
    if seed is not None:
        random.seed(seed)

    if not results:
        return []

    # Group results by stratification field
    groups = defaultdict(list)
    for r in results:
        key = r.get(stratify_by, "unknown")
        groups[key].append(r)

    # Also track by success/error status
    success_results = [r for r in results if r.get("extraction_status") == "success"]
    error_results = [r for r in results if r.get("extraction_status") != "success"]

    sample = []

    # Determine sample allocation
    if stratify_by == 'document_type':
        # Allocate proportionally by document type, but ensure representation
        type_allocations = _calculate_stratified_allocation(
            groups, sample_size, min_per_group=3
        )

        for doc_type, allocation in type_allocations.items():
            type_results = groups[doc_type]

            # Split into success/error within type
            type_success = [r for r in type_results if r.get("extraction_status") == "success"]
            type_errors = [r for r in type_results if r.get("extraction_status") != "success"]

            # Sample from this type
            if include_errors and type_errors:
                # Reserve some slots for errors
                error_slots = min(len(type_errors), max(1, allocation // 3))
                success_slots = allocation - error_slots

                sampled_errors = random.sample(type_errors, min(error_slots, len(type_errors)))
                sampled_success = random.sample(type_success, min(success_slots, len(type_success)))

                sample.extend(sampled_errors)
                sample.extend(sampled_success)
            else:
                sampled = random.sample(type_results, min(allocation, len(type_results)))
                sample.extend(sampled)

    else:
        # Simple stratified sampling by the specified field
        allocations = _calculate_stratified_allocation(groups, sample_size, min_per_group=2)

        for key, allocation in allocations.items():
            group_results = groups[key]
            sampled = random.sample(group_results, min(allocation, len(group_results)))
            sample.extend(sampled)

    # Add sampling metadata
    for i, item in enumerate(sample):
        item["_sampling_metadata"] = {
            "sample_index": i + 1,
            "total_sample_size": len(sample),
            "stratified_by": stratify_by
        }

    return sample


def _calculate_stratified_allocation(groups: dict,
                                      total_size: int,
                                      min_per_group: int = 2) -> dict:
    """Calculate how many samples to take from each group."""
    total_items = sum(len(g) for g in groups.values())

    if total_items == 0:
        return {}

    allocations = {}

    # Proportional allocation
    for key, items in groups.items():
        proportion = len(items) / total_items
        allocation = max(min_per_group, int(proportion * total_size))
        allocations[key] = min(allocation, len(items))

    # Adjust to hit target size
    current_total = sum(allocations.values())
    if current_total < total_size:
        # Add more from largest groups
        remaining = total_size - current_total
        sorted_groups = sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)
        for key, items in sorted_groups:
            if remaining <= 0:
                break
            can_add = len(items) - allocations[key]
            to_add = min(can_add, remaining)
            allocations[key] += to_add
            remaining -= to_add

    return allocations


def generate_validation_worksheet(sample: list[dict],
                                   output_path: str | Path,
                                   max_activities: int = 3) -> bool:
    """
    Generate an Excel worksheet for manual validation.

    Args:
        sample: List of sampled extraction results
        output_path: Path for output Excel file
        max_activities: Maximum number of activity columns to include

    Returns:
        True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return False

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Build headers
    headers = [
        "Sample #",
        "Company Number",
        "Source File",
        "Document Type",
        "Extraction Status",
    ]

    # Add activity columns
    for i in range(1, max_activities + 1):
        headers.extend([
            f"Extracted Activity {i}",
            f"Extracted Benefit {i}",
            f"Correct Activity {i}",
            f"Correct Benefit {i}",
            f"Is Match {i}",
            f"Error Type {i}",
        ])

    headers.extend([
        "Overall Rating",
        "Notes"
    ])

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Write data rows
    for row_idx, result in enumerate(sample, 2):
        col = 1

        # Basic info
        ws.cell(row=row_idx, column=col, value=row_idx - 1).border = thin_border
        col += 1

        ws.cell(row=row_idx, column=col, value=result.get("company_number", "")).border = thin_border
        col += 1

        source_file = result.get("extraction_metadata", {}).get("source_file", "")
        ws.cell(row=row_idx, column=col, value=source_file).border = thin_border
        col += 1

        ws.cell(row=row_idx, column=col, value=result.get("document_type", "")).border = thin_border
        col += 1

        ws.cell(row=row_idx, column=col, value=result.get("extraction_status", "")).border = thin_border
        col += 1

        # Activity columns
        activities = result.get("section_b", {}).get("activities", [])

        for i in range(max_activities):
            if i < len(activities):
                act = activities[i]
                # Extracted activity
                ws.cell(row=row_idx, column=col, value=act.get("activity", "")).border = thin_border
                col += 1
                # Extracted benefit
                ws.cell(row=row_idx, column=col, value=act.get("benefit", "")).border = thin_border
                col += 1
            else:
                # Empty extracted columns
                ws.cell(row=row_idx, column=col, value="").border = thin_border
                col += 1
                ws.cell(row=row_idx, column=col, value="").border = thin_border
                col += 1

            # Input columns (highlighted for human input)
            for _ in range(4):  # Correct Activity, Correct Benefit, Is Match, Error Type
                cell = ws.cell(row=row_idx, column=col, value="")
                cell.fill = input_fill
                cell.border = thin_border
                col += 1

        # Overall rating and notes (input columns)
        cell = ws.cell(row=row_idx, column=col, value="")
        cell.fill = input_fill
        cell.border = thin_border
        col += 1

        cell = ws.cell(row=row_idx, column=col, value="")
        cell.fill = input_fill
        cell.border = thin_border

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Add validation instructions sheet
    instructions = wb.create_sheet("Instructions")
    _add_instructions_sheet(instructions)

    # Add dropdown lists sheet
    dropdowns = wb.create_sheet("Dropdowns")
    _add_dropdown_values(dropdowns)

    # Save
    wb.save(output_path)
    return True


def _add_instructions_sheet(ws):
    """Add instructions for validators."""
    instructions = [
        ("CIC Extraction Validation Instructions", ""),
        ("", ""),
        ("Purpose:", "Validate the accuracy of automated extraction from CIC incorporation documents"),
        ("", ""),
        ("Steps:", ""),
        ("1.", "Open the source PDF file listed in 'Source File' column"),
        ("2.", "Locate Section B: Community Interest Statement - Activities & Related Benefit"),
        ("3.", "Compare extracted text with actual PDF content"),
        ("4.", "Fill in 'Correct Activity' and 'Correct Benefit' columns with actual text"),
        ("5.", "Mark 'Is Match' as: Y (exact match), P (partial match), N (no match)"),
        ("6.", "If not matching, select error type from: missing_data, partial_extraction, ocr_errors, wrong_section, truncated"),
        ("7.", "Set 'Overall Rating': Excellent, Good, Fair, Poor"),
        ("8.", "Add any notes in the Notes column"),
        ("", ""),
        ("Rating Guide:", ""),
        ("Excellent:", "All activities correctly extracted with complete text"),
        ("Good:", "Minor differences (punctuation, spacing) but content correct"),
        ("Fair:", "Some content missing or errors, but main activities captured"),
        ("Poor:", "Significant missing content, wrong section, or garbled text"),
    ]

    for row_idx, (label, value) in enumerate(instructions, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80


def _add_dropdown_values(ws):
    """Add dropdown value lists for data validation."""
    ws.cell(row=1, column=1, value="Is Match Values")
    for i, val in enumerate(["Y", "P", "N"], 2):
        ws.cell(row=i, column=1, value=val)

    ws.cell(row=1, column=2, value="Error Types")
    error_types = ["", "missing_data", "partial_extraction", "ocr_errors", "wrong_section", "truncated", "other"]
    for i, val in enumerate(error_types, 2):
        ws.cell(row=i, column=2, value=val)

    ws.cell(row=1, column=3, value="Overall Rating")
    ratings = ["Excellent", "Good", "Fair", "Poor"]
    for i, val in enumerate(ratings, 2):
        ws.cell(row=i, column=3, value=val)


def load_completed_validation(filepath: str | Path) -> list[dict]:
    """
    Load completed validation worksheet.

    Args:
        filepath: Path to completed Excel file

    Returns:
        List of validation results with human annotations
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Validation"]

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue

        result = dict(zip(headers, row))
        results.append(result)

    return results


def calculate_validation_accuracy(validations: list[dict],
                                   max_activities: int = 3) -> dict:
    """
    Calculate accuracy from completed validation worksheet.

    Args:
        validations: List of validation results from load_completed_validation
        max_activities: Number of activity columns in worksheet

    Returns:
        Accuracy metrics dictionary
    """
    total_activities = 0
    exact_matches = 0
    partial_matches = 0
    no_matches = 0

    rating_counts = defaultdict(int)

    for v in validations:
        # Count overall ratings
        rating = v.get("Overall Rating", "")
        if rating:
            rating_counts[rating] += 1

        # Count activity matches
        for i in range(1, max_activities + 1):
            is_match = v.get(f"Is Match {i}", "")
            if is_match:
                total_activities += 1
                if is_match.upper() == "Y":
                    exact_matches += 1
                elif is_match.upper() == "P":
                    partial_matches += 1
                else:
                    no_matches += 1

    return {
        "total_documents": len(validations),
        "total_activities_validated": total_activities,
        "exact_match_count": exact_matches,
        "partial_match_count": partial_matches,
        "no_match_count": no_matches,
        "exact_match_rate": round(exact_matches / total_activities, 4) if total_activities else 0,
        "partial_or_better_rate": round((exact_matches + partial_matches) / total_activities, 4) if total_activities else 0,
        "rating_distribution": dict(rating_counts)
    }
